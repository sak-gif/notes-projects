import matplotlib.pyplot as plt
import openpyxl as xl
import os

# Handle the case where files might not be found
try:
    S1_WB = xl.load_workbook("files/PV S1-TC_2024-2025_finale.xlsx")
    SHEET_S1 = S1_WB["Worksheet"]

    S2_DSI_WB = xl.load_workbook("files/PV S2-DSI_2024-2025_finale.xlsx")
    SHEET_S2_DSI = S2_DSI_WB["Worksheet"]

    S2_RSS_WB = xl.load_workbook("files/PV S2-RSS_2024-2025_finale.xlsx")
    SHEET_S2_RSS = S2_RSS_WB["Worksheet"]

    S2_DWM_WB = xl.load_workbook("files/PV S2-DWM_2024-2025_finale.xlsx")
    SHEET_S2_DWM = S2_DWM_WB["Worksheet"]
    
    # Mapping of filiere to sheet
    SHEET_MAP = {
        'S1': SHEET_S1,
        'DSI': SHEET_S2_DSI,
        'RSS': SHEET_S2_RSS,
        'DWM': SHEET_S2_DWM,
    }
except FileNotFoundError as e:
    print(f"Error: The file {e.filename} was not found. Please ensure all Excel files are in the correct directory.")
    # Exit the program gracefully if a file is missing
    os._exit(1)

# The rest of your code remains the same as provided, with the following modifications:

def statistique(admis, ajournee):
    """Generates and saves a pie chart of admitted vs. ajourned students."""
    values = [len(admis), len(ajournee)]
    labels = ["Admis", "Ajournés"]
    colors = ["#4CAF50", "#F44336"]
    
    # Create a new figure
    fig, ax = plt.subplots() 
    
    ax.pie(
        values,
        labels=labels,
        colors=colors,
        autopct="%.1f%%",
        startangle=90,
        wedgeprops={"edgecolor": "white"}
    )
    ax.set_title("Répartition des étudiants : Admis vs Ajournés", fontsize=14, fontweight="bold")
    ax.legend(title="Statut", loc="upper right")
    
    # Save the figure and close it by its reference
    fig.savefig("statut.png")
    plt.close(fig) # This is the key change  plt.close(fig) # This is the key change

def statistiqueDeSemesters(semester, filiere):
    """Calculates and plots the general admission statistics for a semester."""
    admis = []
    ajournee = []
    
    sheet_to_use = None
    if semester == "S1":
        sheet_to_use = SHEET_MAP['S1']
        credit_column = 69
    elif semester == "S2":
        sheet_to_use = SHEET_MAP.get(filiere)
        credit_column = 76

    if sheet_to_use:
        for row in range(7, sheet_to_use.max_row + 1):
            cell_value = sheet_to_use.cell(row=row, column=credit_column).value
            
            try:
                credit = float(str(cell_value).replace(",", "."))
                if credit == 30:
                    admis.append(credit)
                else:
                    ajournee.append(credit)
            except (ValueError, TypeError):
                # Skip the row if the value cannot be converted to a number
                continue
        statistique(admis, ajournee)
        return True
    return False

def statistiqueMatiere(semester, filiere, matiere):
    """Calculates and plots admission statistics per subject for a semester/filiere."""
    admis = []
    ajournee = []
    
    sheet_to_use = None
    if semester == "S1":
        sheet_to_use = SHEET_MAP['S1']
    elif semester == "S2":
        sheet_to_use = SHEET_MAP.get(filiere)
        
    if sheet_to_use:
        for row in range(7, sheet_to_use.max_row + 1):
            cell_value = sheet_to_use.cell(row=row, column=matiere).value
            
            try:
                moyenne = float(str(cell_value).replace(",", "."))
                if moyenne >= 10:
                    admis.append(moyenne)
                else:
                    ajournee.append(moyenne)
            except (ValueError, TypeError):
                # Skip the row if the value cannot be converted to a number
                continue
        statistique(admis, ajournee)
        return True
    return False