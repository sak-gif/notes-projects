import os
import openpyxl as xl
import weasyprint as wz
from weasyprint import HTML, CSS

# =================== MODERN CSS WEASYPRINT (PDF_STYLES) ===================
PDF_STYLES = CSS(string='''
@font-face {
    font-family: "Inter";
    src: url("fonts/Inter-Regular.ttf");
}
@font-face {
    font-family: "Inter-Bold";
    src: url("fonts/Inter-Bold.ttf");
    font-weight: bold;
}
@font-face {
    font-family: "JetBrains Mono";
    src: url("fonts/JetBrainsMono-Regular.ttf");
}
@page {
  size: A5;
  margin: 8mm;
}

/* ===== RESET ===== */
* { box-sizing: border-box; }

body {
    font-family: "Inter", sans-serif;
    color: #0f172a;
    font-size: 18pt;
    line-height: 1.6;
    margin: 0;
    padding: 0;
    width: 100%;
    max-width: 100%;
}

.header,
.info-grid,
.department-section,
.subject-cards-wrapper,
.footer {
    width: 100%;
    max-width: 100%;
}

/* ===== HEADER ===== */
.header {
    text-align: center;
    background: #059669;
    padding: 20px;
    border-radius: 12px;
    color: white;
    margin-bottom: 24px;
}
.header h1 { font-size: 20pt; margin: 0; font-weight: bold; }
.header p { margin: 4px 0 0; font-size: 11pt; }

/* ===== INFO GRID ===== */
.info-grid {
    display: grid;
    grid-template-columns: repeat(3, 1fr); /* CHANGED: 4 to 5 columns */
    gap: 12px;
    margin-bottom: 24px;
}
.info-card {
    background: #f9fafb;
    border-radius: 8px;
    padding: 12px;
    text-align: center;
    border: 1px solid #e5e7eb;
}
.info-card strong {
    display: block;
    font-size: 9pt;
    color: #64748b;
    margin-bottom: 6px;
}
.info-card .value { font-size: 12pt; font-weight: bold; }

/* ===== DEPARTMENT SECTIONS ===== */
.department-section {
    background: #ffffff;
    border-radius: 10px;
    padding: 16px;
    margin-bottom: 20px;
    border: 1px solid #e5e7eb;
}
.department-header {
    display: flex;
    justify-content: space-between;
    align-items: center;
    background: #f0fdfa;
    padding: 8px 12px;
    border-radius: 8px;
    margin-bottom: 12px;
    border: 1px solid #d1fae5;
    flex-wrap: wrap;
    gap: 6px;
}
.department-name { font-weight: bold; font-size: 11pt; color: #064e3b; }
.dept-badge {
    background: #06b6d4;
    color: white;
    font-size: 9pt;
    font-weight: bold;
    padding: 4px 8px;
    border-radius: 6px;
}

/* ===== SUBJECT CARDS ===== */
.subject-cards-wrapper {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
    gap: 10px;
}
.subject-card {
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    padding: 10px;
    background: #ffffff;
    font-size: 9pt;
}
.subject-header {
    display: flex;
    justify-content: space-between;
    font-weight: bold;
    margin-bottom: 6px;
    border-bottom: 1px solid #e5e7eb;
    padding-bottom: 4px;
    flex-wrap: wrap;
    gap: 4px;
}
.subject-name { color: #0f172a; font-size: 10pt; }
.subject-grade {
    font-size: 10pt;
    font-weight: bold;
    color: #1d4ed8;
    padding: 2px 6px;
    border: 1px solid #bfdbfe;
    border-radius: 4px;
    background: #dbeafe;
}

/* ===== SUBJECT DETAILS AS CARDS ===== */
.subject-details {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 8px;
    margin-top: 8px;
}
.subject-details .info-card {
    background: #f9fafb;
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    padding: 6px;
    text-align: center;
}
.subject-details .info-card strong {
    display: block;
    font-size: 8pt;
    color: #64748b;
    margin-bottom: 4px;
}
.subject-details .info-card .value {
    font-size: 10pt;
    font-weight: bold;
}

/* ===== BADGES ===== */
.badge {
    padding: 2px 4px;
    border-radius: 6px;
    font-weight: bold;
    font-size: 8pt;
    color: white;
}
.badge.decision-success { background: #16a34a; }
.badge.decision-warning { background: #facc15; color: #111827; }
.badge.decision-danger { background: #dc2626; }

/* ===== FOOTER ===== */
.footer {
    text-align: center;
    font-size: 9pt;
    margin-top: 1cm;
    color: #6b7280;
    padding: 10px;
    border-top: 1px solid #e5e7eb;
}
.footer strong { color: #374151; }

/* ===== RESPONSIVE MEDIA QUERIES ===== */
@media (max-width: 768px) {
    body { font-size: 13pt; }
    .header h1 { font-size: 18pt; }
    .info-grid { grid-template-columns: 1fr 1fr; } /* Added specific rule for this breakpoint */
    .subject-cards-wrapper { grid-template-columns: 1fr; }
}

@media (max-width: 500px) {
    @page {
      size: A6;
      margin: 8mm;
    }
    body {
        font-size: 20pt;
        transform: scale(2);
        padding: 8px;
    }
    .header { padding: 12px; }
    .header h1 { font-size: 16pt; }
    .info-grid { grid-template-columns: 1fr; }
    .subject-cards-wrapper { grid-template-columns: 1fr; }
    .department-header { flex-direction: column; align-items: flex-start; }
    .department-section, .subject-card, .info-card {
        width: 100%;
    }
}
''')


# ===================== UTILITAIRES ET MAPPINGS =====================
def get_decision_class(dec):
    """Retourne la classe CSS selon la décision"""
    # Handle None or empty values gracefully
    dec_str = str(dec).strip().upper() if dec is not None else ""
    if dec_str in ["CI", "AJOURNÉ(E)"]:
        return "decision-warning"
    elif dec_str in ["NC", "NV"]:
        return "decision-danger"
    else:
        return "decision-success"

# ================= EXCEL FILE MAPPINGS =================
S1 = "files/PV S1-TC_2024-2025_finale.xlsx"
S2 = {
    'DSI': 'files/PV S2-DSI_2024-2025_finale.xlsx',
    'RSS': 'files/PV S2-RSS_2024-2025_finale.xlsx',
    'DWM': 'files/PV S2-DWM_2024-2025_finale.xlsx'
}
S3 = {}  # TODO: Add S3 files
S4 = {}  # TODO: Add S4 files
S5 = {}  # TODO: Add S5 files

def find_s_file(matricule, semester_files):
    """Find student in semester files and return specialty and file path"""
    for specialite, file_path in semester_files.items():
        try:
            wb = xl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            for row in range(7, sheet.max_row + 1):
                student_matricule = sheet.cell(row=row, column=2).value
                if str(student_matricule).strip() == str(matricule).strip():
                    return specialite, file_path
        except FileNotFoundError:
            continue
    return None, None

# ================= EXTRACTION FUNCTIONS =================
def extract_student_s1(sheet, matricule):
    """Extract student data for semester 1 — with correct columns and PDF CSS classes"""
    for row in range(7, sheet.max_row + 1):
        student_matricule = sheet.cell(row=row, column=2).value
        if str(student_matricule).strip() == str(matricule).strip():
            nom = f"{sheet.cell(row=row, column=3).value or ''} {sheet.cell(row=row, column=4).value or ''}".strip()
            moyenne = sheet.cell(row=row, column=68).value or "-"
            credits = sheet.cell(row=row, column=69).value or "-"
            decision = sheet.cell(row=row, column=70).value or "-"
            
            matieres = [
                ("Français", 5, 6, 7, 8, 9, "DPR"),
                ("Anglais", 10, 11, 12, 13, 14, "DPR"),
                ("PPP", 15, 16, 17, 18, 19, "DPR"),
                ("Algèbre", 22, 23, 24, 25, 26, "MAI"),
                ("Analyse", 27, 28, 29, 30, 31, "MAI"),
                ("Pix", 32, 33, 34, 35, 36, "MAI"),
                ("C++", 39, 40, 41, 42, 43, "Dev"),
                ("Base de données", 44, 45, 46, 47, 48, "Dev"),
                ("Technology web", 49, 50, 51, 52, 53, "Dev"),
                ("Base info", 56, 57, 58, 59, 60, "SYR"),
                ("Base réseaux", 61, 62, 63, 64, 65, "SYR"),
            ]
            
            dep_data = {
                "DPR": (20, 21),
                "MAI": (37, 38),
                "Dev": (54, 55),
                "SYR": (66, 67),
            }
            
            # group by department
            grouped_by_dept = {}
            for mat, *cols, dept in matieres:
                grouped_by_dept.setdefault(dept, []).append((mat, cols))
            
            departments_html = []
            for dept, items in grouped_by_dept.items():
                moy_dep = sheet.cell(row=row, column=dep_data[dept][0]).value or "-"
                dec_dep = sheet.cell(row=row, column=dep_data[dept][1]).value or "-"
                
                departments_html.append(f"""
<div class="department-section">
    <div class="department-header">
        <div class="department-header-info">
            <span class="department-name">Département {dept}</span>
        </div>
         <div>
        <span class="dept-badge">Moyenne {moy_dep}</span>
        <span class="badge {get_decision_class(dec_dep)}">{dec_dep}</span>
         </div>
    </div>
    <div class="subject-cards-wrapper">""")
                
                for mat, cols in items:
                    d = sheet.cell(row=row, column=cols[0]).value or "-"
                    e = sheet.cell(row=row, column=cols[1]).value or "-"
                    r = sheet.cell(row=row, column=cols[2]).value or "-"
                    m = sheet.cell(row=row, column=cols[3]).value or "-"
                    dec = sheet.cell(row=row, column=cols[4]).value or "-"
                    
                    departments_html.append(f"""
   <div class="subject-card">
        <div class="subject-header">
            <span class="subject-name">{mat}</span>
            <span class="subject-grade">{m}</span>
        </div>
        <div class="subject-details">
            <div class="info-card"><strong>Devoir</strong><div class="value">{d}</div></div>
            <div class="info-card"><strong>Examen</strong><div class="value">{e}</div></div>
            <div class="info-card"><strong>Rattrapage</strong><div class="value">{r}</div></div>
            <div class="info-card"><strong>Décision</strong><div class="value badge {get_decision_class(dec)}">{dec}</div></div>
        </div>
    </div>""")
                
                departments_html.append("""    </div>
</div>""")  # close subject-cards-wrapper + department-section
            
            return {
                "matricule": student_matricule,
                "nom": nom,
                "moyenne": moyenne,
                "credits": credits,
                "decision": decision,
                "departments_html": "\n".join(departments_html)
            }
    return None

def extract_student_s2(sheet, matricule, specialite):
    """Extract student data for semester 2 — with correct columns and PDF CSS classes"""
    for row in range(7, sheet.max_row + 1):
        student_matricule = sheet.cell(row=row, column=2).value
        if str(student_matricule).strip() == str(matricule).strip():
            nom = f"{sheet.cell(row=row, column=3).value or ''} {sheet.cell(row=row, column=4).value or ''}".strip()
            moyenne = sheet.cell(row=row, column=75).value or "-"
            credits = sheet.cell(row=row, column=76).value or "-"
            decision = sheet.cell(row=row, column=77).value or "-"
            
            matieres = [
                ("Français", 5, 6, 7, 8, 9, "DPR"),
                ("Anglais", 10, 11, 12, 13, 14, "DPR"),
                ("PPP", 15, 16, 17, 18, 19, "DPR"),
                ("Python", 22, 23, 24, 25, 26, "Dev"),
                ("Langage web", 27, 28, 29, 30, 31, "Dev"),
                ("Algèbre", 34, 35, 36, 37, 38, "MAI"),
                ("Proba", 39, 40, 41, 42, 43, "MAI"),
                ("Pix", 44, 45, 46, 47, 48, "MAI"),
                ("Système Logique", 51, 52, 53, 54, 55, "SYR"),
                ("Système d'exploitation", 56, 57, 58, 59, 60, "SYR"),
                ("SGBD/Réseaux/CNM", 63, 64, 65, 66, 67, specialite),
                ("Projet Intégrateur", 68, 69, 70, 71, 72, specialite),
            ]
            
            dep_data = {
                "DPR": (20, 21),
                "Dev": (32, 33),
                "MAI": (49, 50),
                "SYR": (61, 62),
                specialite: (73, 74)
            }
            
            grouped_by_dept = {}
            for mat, *cols, dept in matieres:
                grouped_by_dept.setdefault(dept, []).append((mat, cols))
            
            departments_html = []
            for dept, items in grouped_by_dept.items():
                moy_dep = sheet.cell(row=row, column=dep_data[dept][0]).value or "-"
                dec_dep = sheet.cell(row=row, column=dep_data[dept][1]).value or "-"
                
                departments_html.append(f"""
<div class="department-section">
    <div class="department-header">
        <div class="department-header-info">
            <span class="department-name">Département {dept}</span>
        </div>
        <div>
            <span class="dept-badge">Moyenne {moy_dep}</span>
            <span class="badge {get_decision_class(dec_dep)}">{dec_dep}</span>
        </div>
    </div>
    <div class="subject-cards-wrapper">""")
                
                for mat, cols in items:
                    d = sheet.cell(row=row, column=cols[0]).value or "-"
                    e = sheet.cell(row=row, column=cols[1]).value or "-"
                    r = sheet.cell(row=row, column=cols[2]).value or "-"
                    m = sheet.cell(row=row, column=cols[3]).value or "-"
                    dec = sheet.cell(row=row, column=cols[4]).value or "-"
                    
                    departments_html.append(f"""
        <div class="subject-card">
            <div class="subject-header">
                <span class="subject-name">{mat}</span>
                <span class="subject-grade">{m}</span>
            </div>
            <div class="subject-details">
                <div class="info-card"><strong>Devoir</strong><div class="value">{d}</div></div>
                <div class="info-card"><strong>Examen</strong><div class="value">{e}</div></div>
                <div class="info-card"><strong>Rattrapage</strong><div class="value">{r}</div></div>
                <div class="info-card"><strong>Décision</strong><div class="value badge {get_decision_class(dec)}">{dec}</div></div>
            </div>
        </div>""")
                
                departments_html.append("""    </div>
</div>""")
            
            return {
                "matricule": student_matricule,
                "nom": nom,
                "moyenne": moyenne,
                "credits": credits,
                "decision": decision,
                "departments_html": "\n".join(departments_html)
            }
    return None

def generate_html(student, semester_num):
    """Generate complete HTML for the transcript with info-cards"""
    return f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>SUPNUM - Relevé S{semester_num} - {student['nom']}</title>
</head>
<body>
    <div class="header">
        <h1>Relevé de Notes - Semestre {semester_num}</h1>
        <p>SUPNUM - École Supérieure du Numérique • Année Académique 2024-2025</p>
    </div>
    
    <div class="info-grid">
        <div class="info-card">
            <strong>ÉTUDIANT</strong>
            <div class="value">{student['nom']}</div>
        </div>
        <div class="info-card">
            <strong>MATRICULE</strong>
            <div class="value">{student['matricule']}</div>
        </div>
        <div class="info-card">
            <strong>MOYENNE GÉNÉRALE</strong>
            <div class="value">{student['moyenne']}</div>
        </div>
        <div class="info-card">
            <strong>CRÉDITS</strong>
            <div class="value">{student['credits']}</div>
        </div>
        <div class="info-card">
            <strong>DÉCISION</strong>
            <div class="value badge {get_decision_class(student['decision'])}">{student['decision']}</div>
        </div>
    </div>

{student['departments_html']}

    <div class="footer">
        <p><strong>Service Pédagogique 24070 • SUPNUM • École Supérieure du Numérique</strong></p>
        <p>📧 contact@supnum.mr • 📞 +222 24 070 000 • 🌐 www.supnum.mr</p>
    </div>
</body>
</html>
"""

# ================= PDF GENERATION =================
def generate_pdf(student, semester_num, output_dir="pdfs"):
    """Generate and save PDF document"""
    os.makedirs(output_dir, exist_ok=True)
    html_content = generate_html(student, semester_num)
    pdf_path = os.path.join(output_dir, f"releve_S{semester_num}_{student['matricule']}.pdf")
    
    # Important: base_url="." permet de résoudre fonts/... dans le CSS
    HTML(string=html_content, base_url=".").write_pdf(pdf_path, stylesheets=[PDF_STYLES])
    return pdf_path

# ================= PUBLIC FUNCTIONS POUR FLASK =================
def pdf_S1(matricule):
    """Generate PDF for semester 1"""
    try:
        wb = xl.load_workbook(S1, data_only=True)
        sheet = wb.active
        student = extract_student_s1(sheet, matricule)
        if student is None:
            return None
        return generate_pdf(student, 1)
    except Exception as e:
        print(f"Error generating S1 PDF: {e}")
        return None

def pdf_S2(matricule):
    """Generate PDF for semester 2"""
    try:
        specialite, file_path = find_s_file(matricule, S2)
        if file_path is None:
            return None
        wb = xl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        student = extract_student_s2(sheet, matricule, specialite)
        if student is None:
            return None
        return generate_pdf(student, 2)
    except Exception as e:
        print(f"Error generating S2 PDF: {e}")
        return None

# ================= FONCTIONS FUTURES POUR S3-S5 =================
def pdf_S3(matricule):
    """Generate PDF for semester 3 - À implémenter"""
    print("S3 PDF generation not yet implemented")
    return None

def pdf_S4(matricule):
    """Generate PDF for semester 4 - À implémenter"""
    print("S4 PDF generation not yet implemented")
    return None

def pdf_S5(matricule):
    """Generate PDF for semester 5 - À implémenter"""
    print("S5 PDF generation not yet implemented")
    return None

# ================= FONCTION DE TEST =================
def test_pdf_generation():
    """Test function for PDF generation"""
    print("Testing PDF generation...")
    test_matricule = "2024001"  # Remplacez par un vrai matricule pour tester
    
    # Test S1
    result_s1 = pdf_S1(test_matricule)
    if result_s1:
        print(f"S1 PDF generated: {result_s1}")
    else:
        print("S1 PDF generation failed")
    
    # Test S2
    result_s2 = pdf_S2(test_matricule)
    if result_s2:
        print(f"S2 PDF generated: {result_s2}")
    else:
        print("S2 PDF generation failed")

if __name__ == "__main__":
    # Pour tester localement, décommente la ligne suivante :
    test_pdf_generation()
    pass