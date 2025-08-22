import openpyxl as xl

# --- Chargement des fichiers Excel ---
S1 = xl.load_workbook("files/PV S1-TC_2024-2025_finale.xlsx")
SHEET_S1 = S1["Worksheet"]

S2_DSI = xl.load_workbook("files/PV S2-DSI_2024-2025_finale.xlsx")
SHEET_S2_DSI = S2_DSI["Worksheet"]

S2_RSS = xl.load_workbook("files/PV S2-RSS_2024-2025_finale.xlsx")
SHEET_S2_RSS = S2_RSS["Worksheet"]

S2_DWM = xl.load_workbook("files/PV S2-DWM_2024-2025_finale.xlsx")
SHEET_S2_DWM = S2_DWM["Worksheet"]

sheets = [ SHEET_S2_DSI, SHEET_S2_RSS,SHEET_S2_DWM]

# --- Fonctions de tri des moyennes générales ---
def trier_moyenne_s1():
    moyennes = {}
    for row in range(7, SHEET_S1.max_row + 1):
        moyenne_general = SHEET_S1.cell(row=row, column=68).value
        if moyenne_general is not None:
            try:
                moyenne_general = float(str(moyenne_general).replace(",", "."))
                nom = f"{SHEET_S1.cell(row=row, column=3).value or ''} {SHEET_S1.cell(row=row, column=4).value or ''}".strip()
                moyennes[nom] = moyenne_general
            except (ValueError, TypeError):
                continue
    return dict(sorted(moyennes.items(), key=lambda x: x[1], reverse=True))

def dsi():
    moyennes = {}
    for row in range(7, SHEET_S2_DSI.max_row + 1):
        moyenne_general = SHEET_S2_DSI.cell(row=row, column=75).value
        if moyenne_general is not None:
            try:
                moyenne_general = float(str(moyenne_general).replace(",", "."))
                nom = f"{SHEET_S2_DSI.cell(row=row, column=3).value or ''} {SHEET_S2_DSI.cell(row=row, column=4).value or ''}".strip()
                moyennes[nom] = moyenne_general
            except (ValueError, TypeError):
                continue
    return dict(sorted(moyennes.items(), key=lambda x: x[1], reverse=True))

def rss():
    moyennes = {}
    for row in range(7, SHEET_S2_RSS.max_row + 1):
        moyenne_general = SHEET_S2_RSS.cell(row=row, column=75).value
        if moyenne_general is not None:
            try:
                moyenne_general = float(str(moyenne_general).replace(",", "."))
                nom = f"{SHEET_S2_RSS.cell(row=row, column=3).value or ''} {SHEET_S2_RSS.cell(row=row, column=4).value or ''}".strip()
                moyennes[nom] = moyenne_general
            except (ValueError, TypeError):
                continue
    return dict(sorted(moyennes.items(), key=lambda x: x[1], reverse=True))

def dwm():
    moyennes = {}
    for row in range(7, SHEET_S2_DWM.max_row + 1):
        moyenne_general = SHEET_S2_DWM.cell(row=row, column=75).value
        if moyenne_general is not None:
            try:
                moyenne_general = float(str(moyenne_general).replace(",", "."))
                nom = f"{SHEET_S2_DWM.cell(row=row, column=3).value or ''} {SHEET_S2_DWM.cell(row=row, column=4).value or ''}".strip()
                moyennes[nom] = moyenne_general
            except (ValueError, TypeError):
                continue
    return dict(sorted(moyennes.items(), key=lambda x: x[1], reverse=True))

# --- Moyenne par matière ---
def moyenne_matiere(matiere, semester):
    moyenne_du_matiere = {}
    for sheet in sheets:
        
        # Convert matiere to integer if it's a string
        if isinstance(matiere, str):
            try:
                matiere = int(matiere)
            except ValueError:
                return {"error": "matiere must be a valid column number"}

        if semester == "S1":
            for row in range(7, SHEET_S1.max_row + 1):
                moyenne = SHEET_S1.cell(row=row, column=matiere).value
                if moyenne is not None:
                    try:
                        moyenne = float(str(moyenne).replace(",", "."))
                        nom = f"{SHEET_S1.cell(row=row, column=3).value or ''} {SHEET_S1.cell(row=row, column=4).value or ''}".strip()
                        moyenne_du_matiere[nom] = moyenne
                    except (ValueError, TypeError):
                        continue

        elif semester == "S2":
            for row in range(7, sheet.max_row + 1):
                moyenne = sheet.cell(row=row, column=matiere).value
                if moyenne is not None:
                    try:
                        moyenne = float(str(moyenne).replace(",", "."))
                        nom = f"{sheet.cell(row=row, column=3).value or ''} {sheet.cell(row=row, column=4).value or ''}".strip()
                        moyenne_du_matiere[nom] = moyenne
                    except (ValueError, TypeError):
                        continue
        else:
            return {"error": "Invalid semester or filiere"}

        # Retourner les résultats triés
        return dict(sorted(moyenne_du_matiere.items(), key=lambda x: x[1], reverse=True))
    
