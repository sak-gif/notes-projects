import os
import openpyxl as xl
import weasyprint as wz


# ================= MAPPINGS DES FICHIERS EXCEL =================
S1 = "files/PV S1-TC_2024-2025_finale.xlsx"
S2 = {
    'DSI':   'files/PV S2-DSI_2024-2025_finale.xlsx',
    'RSS':   'files/PV S2-RSS_2024-2025_finale.xlsx',
    'DWM':   'files/PV S2-DWM_2024-2025_finale.xlsx'
}
S3 = {} # TODO: À remplir avec les fichiers du semestre 3
S4 = {} # TODO: À remplir avec les fichiers du semestre 4
S5 = {} # TODO: À remplir avec les fichiers du semestre 5

# ================= FONCTIONS UTILITAIRES DE RECHERCHE =================
def find_s_file(matricule, semester_files):
    """
    Parcourt les fichiers d'un semestre donné pour trouver le matricule de l'étudiant.
    Retourne le nom de la spécialité et le chemin du fichier, ou None si non trouvé.
    """
    for specialite, file_path in semester_files.items():
        try:
            wb = xl.load_workbook(file_path, data_only=True)
            sheet = wb.active
            for row in range(7, sheet.max_row + 1):
                student_matricule = sheet.cell(row=row, column=2).value
                if str(student_matricule).strip() == str(matricule).strip():
                    return specialite, file_path
        except FileNotFoundError:
            continue
    return None, None

# ================= FORMATAGE NOTES =================
def format_grade(grade):
    """Retourne un span avec style selon la valeur"""
    try:
        if grade == "-" or grade is None:
            return f'<span class="grade-poor">{grade}</span>'
        
        g = float(grade)
        if g >= 16:
            return f'<span class="grade-excellent">{grade}</span>'
        elif g >= 14:
            return f'<span class="grade-good">{grade}</span>'
        elif g >= 10:
            return f'<span class="grade-average">{grade}</span>'
        elif g >= 5:
            return f'<span class="grade-poor">{grade}</span>'
        else:
            return f'<span class="grade-critical">{grade}</span>'
    except:
        return f'<span class="grade-poor">{grade}</span>'

# ================= EXTRACTION DES NOTES PAR SEMESTRE =================
def extract_student_s1(sheet, matricule):
    """
    Extrait les notes d'un étudiant pour le semestre 1.
    """
    for row in range(7, sheet.max_row + 1):
        student_matricule = sheet.cell(row=row, column=2).value
        if str(student_matricule).strip() == str(matricule).strip():
            nom = f"{sheet.cell(row=row, column=3).value or ''} {sheet.cell(row=row, column=4).value or ''}".strip()
            moyenne = sheet.cell(row=row, column=68).value or "-"
            credits = sheet.cell(row=row, column=69).value or "-"
            decision = sheet.cell(row=row, column=70).value or "-"
            decision_class = "warning" if decision == "Ajourné(e)" else "success"

            matieres = [
                ("Français", 5, 6, 7, 8, 9, "DPR"),
                ("Anglais", 10, 11, 12, 13, 14, "DPR"),
                ("PPP", 15, 16, 17, 18, 19, "DPR"),
                ("Algèbre", 22, 23, 24, 25, 26, "MAI"),
                ("Analyse", 27, 28, 29, 30, 31, "MAI"),
                ("Pix", 32, 33, 34, 35, 36, "MAI"),
                ("C++", 39, 40, 41, 42, 43, "Dev"),
                ("Base de données", 44, 45, 46, 47, 48, "Dev"),
                ("Technology web", 49, 50, 51, 52, 53, "Dev"),
                ("Base info", 56, 57, 58, 59, 60, "SYR"),
                ("Base réseaux", 61, 62, 63, 64, 65, "SYR"),
            ]

            dep_data = {
                "DPR": (20, 21),
                "MAI": (37, 38),
                "Dev": (54, 55),
                "SYR": (66, 67),
            }

            grouped_by_dept = {}
            for mat, *cols, dept in matieres:
                grouped_by_dept.setdefault(dept, []).append((mat, cols))

            departments_html = []
            
            for dept, items in grouped_by_dept.items():
                moy_dep = sheet.cell(row=row, column=dep_data[dept][0]).value or "-"
                dec_dep = sheet.cell(row=row, column=dep_data[dept][1]).value or "-"
                dec_dept_class = "success" if dec_dep == "V" else "danger" if dec_dep == "NV" else "info"

                department_header = f"""
<div class="department-section">
    <div class="department-header">
        <span>Département {dept}</span>
        <div class="department-header-info">
            <strong>Moy. {moy_dep}</strong>
            <span class="dept-badge badge {dec_dept_class}">{dec_dep}</span>
        </div>
    </div>
    <div class="subject-cards-wrapper">
"""
                departments_html.append(department_header)

                for mat, cols in items:
                    d = sheet.cell(row=row, column=cols[0]).value or "-"
                    e = sheet.cell(row=row, column=cols[1]).value or "-"
                    r = sheet.cell(row=row, column=cols[2]).value or "-"
                    m = sheet.cell(row=row, column=cols[3]).value or "-"
                    dec = sheet.cell(row=row, column=cols[4]).value or "-"
                    dec_mat_class = "success" if dec == "C" else "warning" if dec in ("CI", "CE") else "danger" if dec == "NC" else ""
                    
                    card_html = f"""
<div class="subject-card">
    <div class="subject-header">
        <span class="subject-name">{mat}</span>
        <span class="subject-grade">{format_grade(m)}</span>
    </div>
    <div class="subject-details">
        <div class="detail-item">
            <strong>Devoir</strong>
            <span>{format_grade(d)}</span>
        </div>
        <div class="detail-item">
            <strong>Examen</strong>
            <span>{format_grade(e)}</span>
        </div>
        <div class="detail-item">
            <strong>Rattrapage</strong>
            <span>{format_grade(r)}</span>
        </div>
        <div class="detail-item">
            <strong>Décision</strong>
            <span><span class="{dec_mat_class}">{dec}</span></span>
        </div>
    </div>
</div>
""".strip()
                    departments_html.append(card_html)
                
                departments_html.append("</div></div>")
            return {
                "matricule": student_matricule,
                "nom": nom,
                "moyenne": moyenne,
                "credits": credits,
                "decision": decision,
                "decision_class": decision_class,
                "departments_html": "\n".join(departments_html)
            }
    return None

def extract_student_s2(sheet, matricule, specialite):
    """
    Extrait les notes d'un étudiant pour le semestre 2.
    """
    for row in range(7, sheet.max_row + 1):
        student_matricule = sheet.cell(row=row, column=2).value
        if str(student_matricule).strip() == str(matricule).strip():
            nom = f"{sheet.cell(row=row, column=3).value or ''} {sheet.cell(row=row, column=4).value or ''}".strip()
            moyenne = sheet.cell(row=row, column=75).value or "-"
            credits = sheet.cell(row=row, column=76).value or "-"
            decision = sheet.cell(row=row, column=77).value or "-"
            decision_class = "warning" if decision == "Ajourné(e)" else "success"

            matieres = [
                ("Français", 5, 6, 7, 8, 9, "DPR"), 
                ("Anglais", 10, 11, 12, 13, 14, "DPR"),
                ("PPP", 15, 16, 17, 18, 19, "DPR"), 
                ("Python", 22, 23, 24, 25, 26, "Dev"),
                ("Langage web", 27, 28, 29, 30, 31, "Dev"), 
                ("Algèbre", 34, 35, 36, 37, 38, "MAI"),
                ("Proba", 39, 40, 41, 42, 43, "MAI"), 
                ("Pix", 44, 45, 46, 47, 48, "MAI"),
                ("Système Logique", 51, 52, 53, 54, 55, "SYR"), 
                ("Système d'exploitation", 56, 57, 58, 59, 60, "SYR"),
                ("SGBD" if specialite == "DSI" else "Reseaux" if specialite == "RSS" else "CNM", 63, 64, 65, 66, 67, specialite),
                ("Projet Intégrateur", 68, 69, 70, 71, 72, specialite),
            ]
            dep_data = {
                "DPR": (20, 21),
                "Dev": (32, 33),
                "MAI": (49, 50),
                "SYR": (61, 62),
                specialite: (73, 74) 
            }

            grouped_by_dept = {}
            for mat, *cols, dept in matieres:
                grouped_by_dept.setdefault(dept, []).append((mat, cols))

            departments_html = []
            
            for dept, items in grouped_by_dept.items():
                moy_dep = sheet.cell(row=row, column=dep_data[dept][0]).value or "-"
                dec_dep = sheet.cell(row=row, column=dep_data[dept][1]).value or "-"
                dec_dept_class = "success" if dec_dep == "V" else "danger" if dec_dep == "NV" else "info"

                department_header = f"""
<div class="department-section">
    <div class="department-header">
        <span>Département {dept}</span>
        <div class="department-header-info">
            <strong>Moy. {moy_dep}</strong>
            <span class="dept-badge badge {dec_dept_class}">{dec_dep}</span>
        </div>
    </div>
    <div class="subject-cards-wrapper">
"""
                departments_html.append(department_header)

                for mat, cols in items:
                    d = sheet.cell(row=row, column=cols[0]).value or "-"
                    e = sheet.cell(row=row, column=cols[1]).value or "-"
                    r = sheet.cell(row=row, column=cols[2]).value or "-"
                    m = sheet.cell(row=row, column=cols[3]).value or "-"
                    dec = sheet.cell(row=row, column=cols[4]).value or "-"
                    dec_mat_class = "success" if dec == "C" else "warning" if dec in ("CI", "CE") else "danger" if dec == "NC" else ""
                    
                    card_html = f"""
<div class="subject-card">
    <div class="subject-header">
        <span class="subject-name">{mat}</span>
        <span class="subject-grade">{format_grade(m)}</span>
    </div>
    <div class="subject-details">
        <div class="detail-item">
            <strong>Devoir</strong>
            <span>{format_grade(d)}</span>
        </div>
        <div class="detail-item">
            <strong>Examen</strong>
            <span>{format_grade(e)}</span>
        </div>
        <div class="detail-item">
            <strong>Rattrapage</strong>
            <span>{format_grade(r)}</span>
        </div>
        <div class="detail-item">
            <strong>Décision</strong>
            <span><span class="{dec_mat_class}">{dec}</span></span>
        </div>
    </div>
</div>
""".strip()
                    departments_html.append(card_html)
                
                departments_html.append("</div></div>")
            return {
                "matricule": student_matricule,
                "nom": nom,
                "moyenne": moyenne,
                "credits": credits,
                "decision": decision,
                "decision_class": decision_class,
                "departments_html": "\n".join(departments_html)
            }
    return None

def extract_student_s3(sheet, matricule, specialite):
    # TODO: Ajoutez ici la logique d'extraction pour le semestre 3
    # Mettez à jour les listes matieres et dep_data avec les bonnes colonnes et noms
    return None

def extract_student_s4(sheet, matricule, specialite):
    # TODO: Ajoutez ici la logique d'extraction pour le semestre 4
    # Mettez à jour les listes matieres et dep_data avec les bonnes colonnes et noms
    return None
    
def extract_student_s5(sheet, matricule, specialite):
    # TODO: Ajoutez ici la logique d'extraction pour le semestre 5
    # Mettez à jour les listes matieres et dep_data avec les bonnes colonnes et noms
    return None

# ================= GÉNÉRATION HTML =================
def generate_html(student, semester_num):
    """
    Génère la structure HTML complète du relevé de notes.
    """
    return f"""
<div class='body1'>
    <div class="header1">
        <h1>Relevé de Notes - Semestre {semester_num}</h1>
        <p>SUPNUM • Année Académique 2024-2025</p>
    </div>
    <div class="container">
        <div class="info-grid">
            <div class="info-card"><strong>Étudiant</strong><div class="value">{student['nom']}</div></div>
            <div class="info-card"><strong>Matricule</strong><div class="value">{student['matricule']}</div></div>
            <div class="info-card"><strong>Moyenne Générale</strong><div class="value">{format_grade(student['moyenne'])}</div></div>
            <div class="info-card"><strong>Crédits</strong><div class="value">{student['credits']}</div></div>
            <div class="info-card"><strong>Décision</strong><div class="value"><span class="badge {student['decision_class']}">{student['decision']}</span></div></div>
        </div>
        {student['departments_html']}
        <div class="footer1" style="text-align: center; font-size: 0.75rem; color: #64748b; margin-top: 1rem;">
            Service Pédagogique Muhamad Mahmud SAK • SUPNUM • École Supérieure du Numérique<br>
            📧 24070@supnum.mr • 📞 +222 32 16 01 26 • 🌐 www.supnum.mr
        </div>
    </div>
</div>
"""
def sem1(matricule):
    try:
        wb = xl.load_workbook(S1, data_only=True)
        sheet = wb.active
        student = extract_student_s1(sheet, matricule)
        if student is None:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Matricule {matricule} non trouvé.</div>"
        return generate_html(student, 1)
    except Exception as e:
        return f"<div style='color: red; padding: 2rem; text-align:center'>Erreur: {str(e)}</div>"


def sem2(matricule):
    try:
        specialite, file_path = find_s_file(matricule, S2)
        if file_path is None:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Matricule {matricule} non trouvé dans les fichiers du semestre 2.</div>"

        wb = xl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        student = extract_student_s2(sheet, matricule, specialite)

        if student is None:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Matricule {matricule} non trouvé dans la spécialité {specialite}.</div>"
            
        return generate_html(student, 2)
    except Exception as e:
        return f"<div style='color: red; padding: 2rem; text-align:center'>Erreur: {str(e)}</div>"
    

def sem3(matricule):
    try:
        specialite, file_path = find_s_file(matricule, S3)
        if not S3:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Notes du semestre 3 non disponibles.</div>"
        if file_path is None:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Matricule {matricule} non trouvé dans les fichiers du semestre 3.</div>"

        wb = xl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        student = extract_student_s3(sheet, matricule, specialite)
        
        if student is None:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Notes pour le semestre 3 non valides pour le matricule {matricule}.</div>"

        return generate_html(student, 3)
    except Exception as e:
        return f"<div style='color: red; padding: 2rem; text-align:center'>Erreur lors du traitement des notes: {str(e)}</div>"


def sem4(matricule):
    try:
        specialite, file_path = find_s_file(matricule, S4)
        if not S4:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Notes du semestre 4 non disponibles.</div>"
        if file_path is None:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Matricule {matricule} non trouvé dans les fichiers du semestre 4.</div>"

        wb = xl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        student = extract_student_s4(sheet, matricule, specialite)

        if student is None:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Notes pour le semestre 4 non valides pour le matricule {matricule}.</div>"

        return generate_html(student, 4)
    except Exception as e:
        return f"<div style='color: red; padding: 2rem; text-align:center'>Erreur lors du traitement des notes: {str(e)}</div>"

def sem5(matricule):
    try:
        specialite, file_path = find_s_file(matricule, S5)
        if not S5:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Notes du semestre 5 non disponibles.</div>"
        if file_path is None:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Matricule {matricule} non trouvé dans les fichiers du semestre 5.</div>"

        wb = xl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        student = extract_student_s5(sheet, matricule, specialite)

        if student is None:
            return f"<div style='color: red; padding: 2rem; text-align:center'>Notes pour le semestre 5 non valides pour le matricule {matricule}.</div>"

        return generate_html(student, 5)
    except Exception as e:
        return f"<div style='color: red; padding: 2rem; text-align:center'>Erreur lors du traitement des notes: {str(e)}</div>"


# ================= DEBUG =================
if __name__ == "__main__":
    print("🧪 Test du semestre 1 avec matricule 24070:")
    print(sem1("24070")[:500])
    
    print("\n🧪 Test du semestre 2 avec un matricule de DSI (ex: 24200):")
    print(sem2("24200")[:500])
    
    print("\n🧪 Test du semestre 2 avec un matricule de RSS (ex: 24300):")
    print(sem2("24300")[:500])
    
    print("\n🧪 Test du semestre 2 avec un matricule de DWM (ex: 24400):")
    print(sem2("24400")[:500])
    
    # Exemples de tests pour les nouveaux semestres
    # Notez qu'ils renverront des messages d'erreur tant que les fichiers ne sont pas renseignés
    print("\n🧪 Test du semestre 3 avec un matricule inexistant:")
    print(sem3("00000"))
    
    print("\n🧪 Test du semestre 4 avec un matricule inexistant:")
    print(sem4("00000"))
    
    print("\n🧪 Test du semestre 5 avec un matricule inexistant:")
    print(sem5("00000"))
